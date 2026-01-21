"""
Export Service - PDF and Excel generation
"""
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import List, Dict, Optional
import io


class FacultyPDF(FPDF):
    """Custom PDF class for faculty reports"""
    
    def header(self):
        self.set_font('Arial', 'B', 14)
        self.cell(0, 10, 'Faculty Profile Report', 0, 1, 'C')
        self.ln(5)
    
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')
    
    def section_title(self, title):
        self.set_font('Arial', 'B', 12)
        self.set_fill_color(66, 133, 244)
        self.set_text_color(255, 255, 255)
        self.cell(0, 8, title, 0, 1, 'L', True)
        self.set_text_color(0, 0, 0)
        self.ln(2)
    
    def add_field(self, label, value):
        self.set_font('Arial', 'B', 10)
        self.cell(50, 6, f'{label}:', 0, 0)
        self.set_font('Arial', '', 10)
        self.cell(0, 6, str(value) if value else 'N/A', 0, 1)


def generate_faculty_pdf(data: Dict, academic_year: Optional[str] = None) -> bytes:
    """Generate PDF for a single faculty member"""
    pdf = FacultyPDF()
    pdf.add_page()
    
    # Basic Info
    user = data.get('user', {})
    profile = data.get('profile', [{}])[0] if data.get('profile') else {}
    
    pdf.section_title('Basic Information')
    pdf.add_field('Name', f"{profile.get('name_prefix', '')} {user.get('name', '')}")
    pdf.add_field('Employee ID', user.get('employee_id'))
    pdf.add_field('Email', user.get('email'))
    pdf.add_field('Phone', user.get('phone'))
    pdf.add_field('Designation', profile.get('designation'))
    pdf.add_field('Department', profile.get('department'))
    
    if academic_year:
        pdf.add_field('Academic Year Filter', academic_year)
    
    pdf.ln(5)
    
    # Publications
    publications = data.get('publications', [])
    if publications:
        pdf.section_title(f'Publications ({len(publications)})')
        for i, pub in enumerate(publications, 1):
            pdf.set_font('Arial', '', 9)
            pdf.multi_cell(0, 5, f"{i}. {pub.get('title', 'N/A')} - {pub.get('journal_name', '')} ({pub.get('academic_year', '')})")
        pdf.ln(3)
    
    # Awards
    awards = data.get('awards', [])
    if awards:
        pdf.section_title(f'Awards ({len(awards)})')
        for i, award in enumerate(awards, 1):
            pdf.set_font('Arial', '', 9)
            pdf.multi_cell(0, 5, f"{i}. {award.get('title', 'N/A')} - {award.get('awarding_agency', '')} ({award.get('academic_year', '')})")
        pdf.ln(3)
    
    # Patents
    patents = data.get('patents', [])
    if patents:
        pdf.section_title(f'Patents ({len(patents)})')
        for i, patent in enumerate(patents, 1):
            pdf.set_font('Arial', '', 9)
            pdf.multi_cell(0, 5, f"{i}. {patent.get('title', 'N/A')} - Patent No: {patent.get('patent_number', '')} ({patent.get('academic_year', '')})")
        pdf.ln(3)
    
    # Research Projects
    projects = data.get('research_projects', [])
    if projects:
        pdf.section_title(f'Research Projects ({len(projects)})')
        for i, proj in enumerate(projects, 1):
            pdf.set_font('Arial', '', 9)
            pdf.multi_cell(0, 5, f"{i}. {proj.get('title', 'N/A')} - {proj.get('agency', '')} ({proj.get('academic_year', '')})")
        pdf.ln(3)
    
    # Conferences
    conferences = data.get('conferences', [])
    if conferences:
        pdf.section_title(f'Conferences ({len(conferences)})')
        for i, conf in enumerate(conferences, 1):
            pdf.set_font('Arial', '', 9)
            pdf.multi_cell(0, 5, f"{i}. {conf.get('paper_title', 'N/A')} ({conf.get('academic_year', '')})")
    
    return pdf.output(dest='S').encode('latin-1')


def generate_all_faculty_excel(data: List[Dict], academic_year: Optional[str] = None) -> bytes:
    """Generate Excel file with all faculty summary"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty Summary"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4285F4", end_color="4285F4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Faculty Summary Report - {academic_year or 'All Years'}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ["Name", "Email", "Employee ID", "Designation", "Department", "Publications", "Awards", "Patents"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_num, faculty in enumerate(data, 4):
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=faculty.get(key, ""))
            cell.border = thin_border
    
    # Adjust column widths
    column_widths = [25, 30, 15, 20, 25, 12, 10, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_all_faculty_summary_pdf(
    data: List[Dict], 
    academic_year: Optional[str] = None,
    department: Optional[str] = None
) -> bytes:
    """Generate PDF summary of all faculty"""
    pdf = FPDF()
    pdf.add_page()
    
    # Title
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Faculty Directory Summary', 0, 1, 'C')
    
    pdf.set_font('Arial', '', 10)
    if academic_year:
        pdf.cell(0, 6, f'Academic Year: {academic_year}', 0, 1, 'C')
    if department:
        pdf.cell(0, 6, f'Department: {department}', 0, 1, 'C')
    
    pdf.ln(10)
    
    # Table header
    pdf.set_font('Arial', 'B', 10)
    pdf.set_fill_color(66, 133, 244)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(10, 8, '#', 1, 0, 'C', True)
    pdf.cell(50, 8, 'Name', 1, 0, 'C', True)
    pdf.cell(60, 8, 'Email', 1, 0, 'C', True)
    pdf.cell(30, 8, 'Emp ID', 1, 0, 'C', True)
    pdf.cell(40, 8, 'Department', 1, 1, 'C', True)
    
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Arial', '', 9)
    
    for i, faculty in enumerate(data, 1):
        pdf.cell(10, 7, str(i), 1, 0, 'C')
        pdf.cell(50, 7, faculty.get('name', '')[:25], 1, 0, 'L')
        pdf.cell(60, 7, faculty.get('email', '')[:30], 1, 0, 'L')
        pdf.cell(30, 7, faculty.get('employee_id', ''), 1, 0, 'C')
        pdf.cell(40, 7, (faculty.get('department', '') or '')[:20], 1, 1, 'L')
    
    pdf.ln(10)
    pdf.set_font('Arial', 'I', 10)
    pdf.cell(0, 10, f'Total Faculty: {len(data)}', 0, 1, 'R')
    
    return pdf.output(dest='S').encode('latin-1')
